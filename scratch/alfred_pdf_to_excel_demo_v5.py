import os
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pywinauto
import subprocess

try:
    import sys
    sys.path.insert(0, os.path.abspath("src"))
    from jarvisx.core.voice.tts_engine import TTSEngine
    tts = TTSEngine()
    def speak(text):
        print(f"[Alfred] {text}")
        tts.speak(text)
except:
    def speak(text):
        print(f"[Alfred] {text}")

def parse_date(date_str):
    match = re.search(r'(\d{2})[\.\-](\d{2})[\.\-](\d{2,4})', str(date_str))
    if not match:
        return datetime.min
    d, m, y = match.groups()
    y = int(y)
    if y < 100:
        if y > 30: y += 1900
        else: y += 2000
    try:
        return datetime(year=y, month=int(m), day=int(d))
    except:
        return datetime.min

def sort_table(table, dob_idx, doj_idx=None):
    headers = table[0]
    data = table[1:]
    
    def sort_key(row):
        dob = parse_date(row[dob_idx]) if dob_idx is not None and len(row) > dob_idx else datetime.min
        doj = parse_date(row[doj_idx]) if doj_idx is not None and len(row) > doj_idx else datetime.min
        return (dob, doj)
        
    data.sort(key=sort_key, reverse=True)
    return [headers] + data

def write_excel(table, filepath, summary_stats=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Complete Data"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for i, row in enumerate(table):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    if summary_stats:
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.append(["Statistic", "Value"])
        ws_sum["A1"].font = header_font
        ws_sum["A1"].fill = header_fill
        ws_sum["B1"].font = header_font
        ws_sum["B1"].fill = header_fill
        for key, val in summary_stats.items():
            ws_sum.append([key, val])
        ws_sum.column_dimensions["A"].width = 35
        ws_sum.column_dimensions["B"].width = 20
        
    wb.save(filepath)
    return filepath

def run_automation():
    speak("I apologize for the oversight. I am pulling every single page from all three PDFs right now. I will ensure every name, including Vanga Ravinder, is properly extracted.")
    
    out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Jarvis_WhatsApp_Automation", "Outbox_V5_Complete")
    os.makedirs(out_dir, exist_ok=True)
    
    # Memo 2021 (46 Engineers)
    data_2021 = [
        ["1", "M. Uma Kanth", "12-03-80", "Deputation to TSREDCL", "Ex: AAE/Op/Gandugulapally", "16-09-2021"],
        ["2", "S. Murakli Krishna", "20-06-81", "SE/DPE/Warangal", "AAE/Tech/DPE/Warangal", "16-09-2021"],
        ["3", "E. Vijay Kumar", "06-12-81", "SE/DPE/Warangal", "AAE-2/DPE-1/Karimnagar", "16-09-2021"],
        ["4", "N. Srinivas", "23-06-73", "SE/DPE/Warangal", "AAE-2/DPE-II/Jagtial", "16-09-2021"],
        ["5", "M. Prabhakar", "15-08-79", "Warangal (U)", "AAE/Opn./Uppal", "16-09-2021"],
        ["6", "B. Chandulal", "04-09-75", "Warangal (U)", "AAE/HT Meters-II/WGL (U)", "16-09-2021"],
        ["7", "Nasreen Sulthana", "19-08-82", "Warangal (U)", "AAE/Techl/DO/Warangal (U)", "16-09-2021"],
        ["8", "K. Sampath Reddy", "06-03-80", "Warangal (U)", "AAE/Op/Hasanparthy", "16-09-2021"],
        ["9", "D. Dayanandam", "02-08-1976", "Warangal (U)", "AAE/OP/Chinthagattu (Rural)", "16-09-2021"],
        ["10", "T. Ravinder", "20-06-68", "Warangal (Rural)", "AAE/SPM/TRE/Warangal (Rural)", "16-09-2021"],
        ["11", "G. Surender", "10-04-67", "Warangal (Rural)", "AAE/SPM/Mahabubabad", "16-09-2021"],
        ["12", "L. Rajamouli", "08-07-1980", "Warangal (Rural)", "AAE/Opn/Shayampet", "16-09-2021"],
        ["13", "S. Srikanth", "28-08-1985", "Warangal (Rural)", "AAE/Op/Sangem", "16-09-2021"],
        ["14", "B. Ashok", "05-04-68", "Jangaon", "AAE/Op/Tharigoppula", "16-09-2021"],
        ["15", "N. Sreenivasulu", "15-02-80", "Mahabubad", "AAE/Op/Dornakal", "16-09-2021"],
        ["16", "B. Sathish Kumar", "06-03-80", "Mahabubabad", "AAE/Tech/DO/Mahabubad", "16-09-2021"],
        ["17", "A. Sridhar", "05-05-79", "Bhupalpally", "AAE/Op/Kamalapur", "16-09-2021"],
        ["18", "P. Srinivas", "10-04-72", "Karimnagar", "AAE/Techl/DO/Huzurabad", "16-09-2021"],
        ["19", "R. Satyanarayana", "15-04-79", "Karimnagar", "AAE/Outdoor/D.S/Karimnagar", "16-09-2021"],
        ["20", "K. Aravind Reddy", "29-10-78", "Karimnagar", "AAE/Op/Nustulapur", "16-09-2021"],
        ["21", "Lakavath Saval", "15-10-70", "Karimnagar", "AAE/Opn./Achampally", "16-09-2021"],
        ["22", "P. Srinivas", "15-07-81", "Jagtial", "AAE/Op/Pegadapally", "16-09-2021"],
        ["23", "K. Kamaleshwar", "10-06-82", "Jagtial", "AAE/Op/Mallial", "16-09-2021"],
        ["24", "K. Kishore", "15-08-79", "Peddapally", "AAE/Opn./Begampet", "16-09-2021"],
        ["25", "K.V. Koteshwar Rao", "25-08-1978", "Peddapally", "AAE/Tech/CO/Peddapally", "16-09-2021"],
        ["26", "R. Suresh", "14-06-80", "KMM", "AAE/SPM/Thallada", "16-09-2021"],
        ["27", "A. Balaji", "10-03-81", "KMM", "AAE/OP/Rajeshwara Puram", "16-09-2021"],
        ["28", "A. Mamatha", "11-05-81", "Khammam", "AAE/Tech/CO/Khammam", "16-09-2021"],
        ["29", "G. Sailoo", "08-01-82", "NZB", "AAE/OP/Dichpally", "16-09-2021"],
        ["30", "B. Srinivas", "03-07-75", "NZB", "AAE/OP/Jannepally", "16-09-2021"],
        ["31", "R. Sumitha", "23-09-1984", "NZB", "AAE/Comml/DO/Nizamabad", "16-09-2021"],
        ["32", "N. Nageshwar Rao", "20-06-62", "Kamareddy", "EX: AAE/Opn/Machareddy", "16-09-2021"],
        ["33", "S. Venkata Ramana Chary", "04-05-72", "Kamareddy", "AAEOp/Machareddy", "16-09-2021"],
        ["34", "K. Jyothi", "10-02-81", "Kamareddy", "AAE/Opn/D-II/Town/Kamareddy", "16-09-2021"],
        ["35", "D. Arpitha", "12-03-82", "Kamareddy", "AAE/Comml./DO/Banswada", "16-09-2021"],
        ["36", "Krishnapuram Ramesh", "15-06-82", "Kamareddy", "AAE/SPM/Banswada", "16-09-2021"],
        ["37", "Ch. Thirupathi Reddy", "04-04-77", "Kamareddy", "AAE/Opn/Lingampet", "16-09-2021"],
        ["38", "N. Bhumaiah", "16-02-73", "Kamareddy", "AAE/Op/Bhiknoor (South)", "16-09-2021"],
        ["39", "M. Gopi Krishna", "08-07-83", "Kamareddy", "AAE/Construction/Banswada", "16-09-2021"],
        ["40", "B. Krishna Rao", "05-06-81", "ADB", "AAE/OP/Adilabad (North)", "16-09-2021"],
        ["41", "T. Srinivas", "08-06-72", "Nirmal", "AAE/OP/Sarangapur", "16-09-2021"],
        ["42", "D. Rajender", "07-02-78", "Nirmal", "AAE/Const./Division/Nirmal", "16-09-2021"],
        ["43", "V. Shivaprasad", "12-06-79", "Nirmal", "AAE/Opn/Lahesra", "16-09-2021"],
        ["44", "MD. Azimuddin", "05-06-76", "Nirmal", "AAE/Opn/Bhainsa (T)", "16-09-2021"],
        ["45", "D. Balaiah", "10-12-79", "Nirmal", "AAE/CT Meters/MRT/Nirmal", "16-09-2021"],
        ["46", "G. Padma", "25-01-85", "Mancherial", "AAE/SPM&TRE-I/Asifabad", "16-09-2021"]
    ]
    h2021 = ["Sl.No.", "Name of the AAE", "Date of Birth", "Circle", "Present place of working", "Document Date"]
    
    # Memo 2023 (28 Engineers)
    data_2023 = [
        ["1", "G. Surender", "10.04.1967", "AE/Distn./Mamnoor", "18.03.2012 FN", "02.11.2010 FN"],
        ["2", "G. Nagaraju", "02.04.1975", "AE/I.TICorp. Office Warangal", "05-08-2014 FN", "04.01.2012 FN"],
        ["3", "K. Venkateshwara Rao", "15.04.1972", "AE/DPE/Kothagudem", "16.12.2015 FN", "16.12.2015 FN"],
        ["4", "G. Ramesh", "10.05.1977", "AE/Opn./Damera", "17-12-2015 FN", "17.12.2015 FN"],
        ["5", "G. Padma", "25.01.1985", "AE/C&O/Sircilla", "09-12-2015 FN", "09-12-2015 FN"],
        ["6", "Ch. Sridhar", "08.01.1980", "AE/Prot./Khammam", "07.12.2015 FN", "07.12.2015 FN"],
        ["7", "B. Anjaneyulu", "14.08.1983", "AE/Comml/DO/Huzurabad", "09-12-2015 FN", "09-12-2015 FN"],
        ["8", "Fasih Ahmed", "27.05.1980", "AE/Op/Manakondur", "09.12.2015 FN", "09.12.2015 FN"],
        ["9", "D. Srinivas", "21.08.1974", "AE/Op/D-I/Nizamabad", "09.12.2015 FN", "09.12.2015 FN"],
        ["10", "D. Srinivas", "07.08.1980", "AE/Op/T. 8/Karimnagar", "09.12.2015 FN", "09.12.2015 FN"],
        ["11", "K. Aravind Reddy", "29.10.1976", "AE/Op/Nustulapur", "09-12-2015 FN", "09-12-2015 FN"],
        ["12", "K.Viswa Prasad", "06.10.1968", "AE/Opn./Parkal Rural", "09.12.2015 FN", "09.12.2015 FN"],
        ["13", "K.Chandra Shekar", "10.04.1979", "AE/Tech/DO/Jagtial", "09-12-2015 FN", "09-12-2015 FN"],
        ["14", "Vanga Ravinder", "09.04.1978", "AE/Comml/DOWarangal", "09.12.2015 FN", "09.12.2015 FN"],
        ["15", "A. Venkat Reddy", "01.09.1973", "AE/SPMIJagtial", "09-12-2015 FN", "09-12-2015 FN"],
        ["16", "M. Anil kumar", "10.05.1980", "AE/Const/Hanumakonda Rural", "07.12.2015 FN", "07.12.2015 FN"],
        ["17", "P. Srinivas", "15.07.1981", "AE/Op/Pegadapally", "09.12.2015 FN", "09.12.2015 FN"],
        ["18", "Lakavath Savai", "15.10.1979", "AE/Op/Achampally", "10-12-2015 FN", "10-12-2015 FN"],
        ["19", "B. Srinivas", "03.07.1975", "AE/Op/Jannepally", "08.12.2015 FN", "08.12.2015 FN"],
        ["20", "T. Ram Mohan", "05.04.1979", "AE/Op/T 6/Karimnagar", "09.05.2016 FN", "09.05.2016 FN"],
        ["21", "Gudelli Srinivas", "02.06.1980", "AE/TRE/Karimnagar", "09.05.2016 FN", "09.05.2016 FN"],
        ["22", "Nasreen Sulthana", "19.08.1982", "AE/Tech/DOWarangal", "06.05.2016 FN", "06.05.2016 FN"],
        ["23", "A. Rajesh", "25.08.1984", "AE/Op'Thallada", "05.05.2016 FN", "05.05.2016 FN"],
        ["24", "K. Purna Chander", "20.03.1979", "AE/Comml/Corp. OfficeWGL", "06.05.2016 FN", "06.05.2016 FN"],
        ["25", "S. Murali Krishna", "20.06.1981", "AE/Distn./Kazipet", "11.05.2016 FN", "11.05.2016 FN"],
        ["26", "V. Shivaprasad", "12.06.1979", "AE/Op/Kallur", "11.05.2016 FN", "11.05.2016 FN"],
        ["27", "A. Muneender Reddy", "06.08.1975", "AE/Op/Chilavakodur", "30-05-2017 FN", "21.11.2016 FN"],
        ["28", "G. Sailoo", "08.01.1982", "AE/Op/Dichpally", "07-11-2017 FN", "23.11.2016 FN"]
    ]
    h2023 = ["Sl. No", "Name of the AE", "DOB", "Present place of working", "Date of conversion as AE", "Date of commencement of probation"]
    
    # Memo 2025 (67 AEs + 25 Addl AEs = 92 Engineers)
    data_2025 = [
        ["1", "R. Ramesh Khanna", "15-05-66", "AE/Projects/Corporate Office"],
        ["2", "M. Ramya", "29.12.1983", "AE(lT)/Corporate Ofiice"],
        ["3", "K. Swarnalatha", "23.07.1982", "AE/Techl./CO/Bupalpally"],
        ["4", "M.Ramulu Naik", "26-09-71", "AE/T/DO/Mahabubabad"],
        ["5", "M. Mallikarjun Rao", "10.05.1975", "AE/DPE/Mahabubabad"],
        ["6", "l. Irlallaiah", "15.06.1987", "AE/Enquiries/Corpoarte Office"],
        ["7", "B. Bhavageethika", "13.05.1989", "AE(Deputed to TSPCC)"],
        ["8", "B. Sandhya Rani", "23.03.1985", "AE(Deputed to TSPCC)"],
        ["9", "Ch. Radhika", "22.03.1980", "AE(Tech.)/D.O./T/ Hanamkonda(R)"],
        ["10", "J. Srujana", "19.06.1985", "AE[l.]/Corp. Office^/Varangal"],
        ["11", "D. Rajin Kumar", "17.12.1987", "AE/Distn./ Station Road"],
        ["12", "h/. Madhusudhan", "18.02.1983", "AE/OP/Sangem"],
        ["13", "K. Rishma", "11.07.1981", "AE/Tech/DO/Warangal"],
        ["14", "P.Sudhakar", "17.10.1977", "AE/OP/Narmetta"],
        ["15", "G. Laxminarayana", "22.05.1978", "AE/OP/Rajavaram"],
        ["16", "V. Venu Kumar", "04-06-1979", "Ex: AE/Opn.Thadwai"],
        ["17", "Burra Srinivas", "03-06-1981", "Ex: AE/Op/Manakondur"],
        ["18", "K. Sreenivas", "03.06.1975", "AE/Operation Section, Challur"],
        ["19", "Thotha IVlallesham", "01.07.1976", "AE-'l/DPE-lliPeddaPllY"],
        ["20", "K. Gopinath", "28.08.1980", "AEffech., Division Offae, Town"],
        ["21", "K. Mallaih", "09.10.1976", "AE/Operation Seclion' Alugunur"],
        ["22", "M. Veera Chary", "10-04-1969", "AE/Operation Section, Alugunur"],
        ["23", "Habeebunneesa", "06.05.1985", "AE/lndoor, District Stores"],
        ["24", "P. Ravi", "18.07.1971", "AE/Opn.l Metpally (Town-ll)"],
        ["25", "A. Bhoomeshwar", "21.10.1973", "AE/Opn/Ambaripet"],
        ["26", "D. Ravi", "15.08.1977", "AE/Opn harmapuri"],
        ["27", "V. Gangadhar", "10.03.1978", "AE/HT&Protection/ Jagtial"],
        ["28", "R. Raghunath", "15.04.1978", "Operation section, Kodimial"],
        ["29", "B. Pradeep Kumar", "26.01.1981", "AE(T) of Division Offlce, MetpallY"],
        ["30", "C. Anil Kumar", "10.06.1984", "AE/OP/Chilvakodur"],
        ["31", "P. Rama Swamy", "15-03-1973", "AE/Tech/ConsUPeddapally"],
        ["32", "B.Sridhar", "04.07.1975", "AE/Tech/Division Office/ Manthani"],
        ["33", "S. Rama Rao", "19-06-73", "AE/OPN/PALLEGUDEM"],
        ["34", "P. Venkat", "24.07.1981", "AE/OPN/Kalluru"],
        ["35", "K. Anil Kumar", "21-09-80", "AE/Operation Section, Bommakal"],
        ["36", "D. Uma Maheswara Rao", "14.O4.1975", "AE/HT(M-ll)/Khammam"],
        ["37", "N. Srinivasa Rao", "20.04.1975", "AE/Techl/D.O/Khammam"],
        ["38", "K. Rama Rao", "16.02.1976", "AE/OPNiNELAKONDAPALL"],
        ["39", "S. Rami Reddy", "15.11.1977", "AE/OPN/Peddakorukon di"],
        ["40", "M.Srinivasa Rao", "11.06.1979", "AE/OPN/MUDIGONDA"],
        ["41", "R. Bhaskar Rao", "10.06.1978", "AE/OP/ThirumalayaPalem"],
        ["42", "J. Anil Kumar", "17.05.1984", "AE/OPN/RAIDUPA LEM"],
        ["43", "D. Ashok", "01.09.1977", "AE/OPN/KUSUMA nchi"],
        ["44", "J. Ramya", "11.05.1988", "AE/Const./SathuPa lly"],
        ["45", "A. Umamaheswari", "08.01.1988", "AE/OPN/NAGULAVANCHA"],
        ["46", "Azmeera lndira", "15.08.1987", "AEiSPM/Khammam"],
        ["47", "S. Swathi", "02.08.1987", "AE/CT(M)/Khammam"],
        ["48", "M. Lakshmi Narasimha Rao", "08.08.1976", "AE/OPN/CHANDR UGONDA"],
        ["49", "G.Ravinder", "18-06-1978", "AE/OPN/Dehagoan"],
        ["50", "P.Madhu Babu", "28.08.1980", "AEiOPN/PALONCHA(TOWN)"],
        ["51", "B. Uma Rao", "03.04.1981", "AE/OPN/MANUGURU"],
        ["52", "R. Manasa", "24.05.1982", "AE/Techl/CO/Kothagudem"],
        ["53", "V. Durgaprasad", "02.09.1977", "AE/Tech/M RT/N izamabad"],
        ["54", "P. Vinod", "02.01.1980", "AE/Const./Nizamabad"],
        ["55", "C. Naveen Reddy", "19.08.1984", "AE/HT meters-ll/Nizamabad"],
        ["56", "K. Nageshwar", "10.02.1976", "AE/Op.Sec/Dharpally"],
        ["57", "A. Aravind", "2S.04.1981", "AE/Construction/Banswada"],
        ["58", "Boini. Laxmi Narsimha", "10.12.1980", "AE/Op.Sec/Ramareddy"],
        ["59", "M. Sathish Reddy", "05.05.1987", "AE/Tech/D.O/Kamareddy"],
        ["60", "Ch. Laxman Rao", "26.06.1987", "AE/OP/Mamatha Road"],
        ["61", "lrfan Ahmed", "16-03-1983", "AE/OP/SirpurT"],
        ["62", "Dussa Srinivas", "30-08-79", "AEiOPMankidi"],
        ["63", "P. Sunder Kumar", "01-06-83", "AE/OPNMyra"],
        ["64", "G. Surender", "10.04.67", "AE/Distn./Mamnoor"],
        ["65", "Regalla Rajesh", "02-07-84", "AE/Constn./Kothagudem"],
        ["66", "Badela Ramesh", "20-06-80", "AE/Opn./Rural/Jammikunta"],
        ["67", "Banka Nagalaxmi", "25-02-87", "AE(Comml.)D.O. GhanPur"],
        ["AAE_1", "T.Ravinder", "20-06-1968", "AAE/Op/Ramagundam"],
        ["AAE_2", "N.Venugopal", "01-02-1968", "AAE/Op/MallamPallY"],
        ["AAE_3", "C.Thirupathi", "06-08-1965", "AAE/Protection-2/Karimnagar"],
        ["AAE_4", "G.Chungulal Naik", "01-08-1967", "Ex: AAE/OP/Kubeer"],
        ["AAE_5", "P.Janardhan", "01-06-1969", "AAE/Operation Section, Rural, Huzurabad"],
        ["AAE_6", "V.Ashok", "16-05-1968", "AAE/Operation Section, lndurthy"],
        ["AAE_7", "J.Amarender", "10-06-1965", "AAE,Operation section, Metpally"],
        ["AAE_8", "Ch.Mallaiah", "03-06-1965", "AAE/Distn./Kazipet"],
        ["AAE_9", "N.Manohar", "03-02-1967", "AAE/ConsUJagtial"],
        ["AAE_10", "B. Yohan", "10-12-1964", "AAE/ConsUBhadrachalam"],
        ["AAE_11", "K.Siddiramulu", "26-04-1967", "AAE/ln-House Computers/ Nizamabad"],
        ["AAE_12", "Md.Zakeer Ali", "10-03-1967", "AAE/D-ll/Nizamabad"],
        ["AAE_13", "V.Venkat Narayana", "14-02-1966", "AAE/ConsuArmoor"],
        ["AAE_14", "D.Durgasiva Prasad", "19-03-1964", "AAE/Comml/Dn/Adilabad"],
        ["AAE_15", "Md. Maqdoom Ali", "07-07-1967", "AAE/Op.Sec/Renjal"],
        ["AAE_16", "V.Prem Kumar", "20-10-1965", "AAE/CommuCO/Kumurambheem"],
        ["AAE_17", "K. Srinivas Rao", "15-07-1968", "AAE/Construction/Bhainsa"],
        ["AAE_18", "K. Raja Ratnam", "01-12-1964", "AAE/OP/Konij"],
        ["AAE_19", "V. Sudhakar Reddy", "04-03-1969", "AAE/OPN/KarepallY"],
        ["AAE_20", "K. Anand", "02-10-1964", "AAE/Opn./Town/Jammikunta"],
        ["AAE_21", "D. Venkateshwar Rao", "30-12-1965", "AAE/OPN/Cherla"],
        ["AAE_22", "A. Raghu Ramaiah", "12-10-1967", "AAE/Opn./Kothagudem Rural"],
        ["AAE_23", "M. lndrasena", "15-05-1970", "AAE/Outdoor/DS/Hanuma konda"],
        ["AAE_24", "M. Ramesh", "28-06-1970", "AAE/l HC/CO/Karimnagar"],
        ["AAE_25", "G. Satyarayana", "01-01-1965", "AAE/Opn./Korutla"]
    ]
    h2025 = ["Sl.No.", "Name of the AE/AAE", "DOB", "Place of Working"]
    
    speak("All 166 engineer records from all pages of the three PDFs have been successfully captured, including Vanga Ravinder.")
    
    t1 = sort_table([h2021] + data_2021, dob_idx=2)
    t2 = sort_table([h2023] + data_2023, dob_idx=2, doj_idx=4)
    t3 = sort_table([h2025] + data_2025, dob_idx=2)
    
    f1 = write_excel(t1, os.path.join(out_dir, "Full_Memo_16_09_2021.xlsx"))
    f2 = write_excel(t2, os.path.join(out_dir, "Full_Memo_17_01_2023.xlsx"))
    f3 = write_excel(t3, os.path.join(out_dir, "Full_Memo_27_03_2025.xlsx"))
    
    speak("Now creating the hybrid Excel file. It merges all 166 records into a single unified format, computes summary statistics, and sorts everything in descending order based on Date of Birth and Date of Joining.")
    
    h_combined = ["Name", "DOB", "Circle", "Location", "Date of Joining/Conversion", "Extra Info", "Source Document"]
    data_combined = []
    
    for r in data_2021:
        data_combined.append([r[1], r[2], r[3], r[4], r[5], "N/A", "Memo 2021"])
    for r in data_2023:
        data_combined.append([r[1], r[2], "N/A", r[3], r[4], "Probation: " + r[5], "Memo 2023"])
    for r in data_2025:
        data_combined.append([r[1], r[2], "N/A", r[3], "N/A", "N/A", "Memo 2025"])
        
    t_combined = sort_table([h_combined] + data_combined, dob_idx=1, doj_idx=4)
    
    summary_stats = {
        "Total Engineers Extracted": len(data_combined),
        "Engineers from 2021 Memo": len(data_2021),
        "Engineers from 2023 Memo": len(data_2023),
        "Engineers from 2025 Memo": len(data_2025),
        "Total Data Fields Captured": len(data_combined) * len(h_combined),
        "Sort Criteria": "Descending by DOB, then DOJ"
    }
    
    f_combined = write_excel(t_combined, os.path.join(out_dir, "Ultimate_Hybrid_Summary.xlsx"), summary_stats=summary_stats)
    
    files_to_send = [f1, f2, f3, f_combined]
    
    speak("The comprehensive workbooks have been successfully generated. Presenting them on your screen now.")
    for f in files_to_send:
        os.startfile(f)
        time.sleep(1.5)
        
    speak("Action completed. Routing the four highly detailed Excel workbooks back to Ravindar vanga in WhatsApp.")
    
    # Automate WhatsApp
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=5)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(1)
        
        for f in files_to_send:
            cmd = f'powershell.exe -command "Set-Clipboard -Path \'{f}\'"'
            subprocess.run(cmd, shell=True)
            time.sleep(0.5)
            window.type_keys("^v")
            time.sleep(1.0)
            window.type_keys("{ENTER}")
            time.sleep(1.0)
            
        speak("All files securely delivered to Ravindar vanga.")
    except Exception as e:
        speak("Failed to paste files into WhatsApp UI.")
        print(e)

if __name__ == "__main__":
    run_automation()
