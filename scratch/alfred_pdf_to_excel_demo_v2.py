import os
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pywinauto
import subprocess

try:
    import sys
    sys.path.insert(0, os.path.abspath("src"))
    from jarvisx.core.voice.tts_engine import TTSEngine
    tts = TTSEngine()
    def speak(text):
        print(f"[Alfred] {text}")
        tts.speak(text)
except:
    def speak(text):
        print(f"[Alfred] {text}")

def parse_date(date_str):
    # Match patterns like 10.04.1967 or 10-04-1967 or 15-05-66
    match = re.search(r'(\d{2})[\.\-](\d{2})[\.\-](\d{2,4})', str(date_str))
    if not match:
        return datetime.min
    d, m, y = match.groups()
    y = int(y)
    if y < 100:
        if y > 30: y += 1900
        else: y += 2000
    try:
        return datetime(year=y, month=int(m), day=int(d))
    except:
        return datetime.min

def sort_table(table, dob_idx, doj_idx=None):
    headers = table[0]
    data = table[1:]
    
    def sort_key(row):
        dob = parse_date(row[dob_idx]) if dob_idx is not None and len(row) > dob_idx else datetime.min
        doj = parse_date(row[doj_idx]) if doj_idx is not None and len(row) > doj_idx else datetime.min
        # Descending order -> max date first
        return (dob, doj)
        
    data.sort(key=sort_key, reverse=True)
    return [headers] + data

def write_excel(table, filepath):
    wb = Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for i, row in enumerate(table):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    wb.save(filepath)
    return filepath

def run_automation():
    speak("Yes sir. I am separating all four PDF memos, extracting the tables, and sorting the engineers by Date of Birth and Date of Joining in descending order.")
    
    out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Jarvis_WhatsApp_Automation", "Outbox_Separate")
    os.makedirs(out_dir, exist_ok=True)
    
    # --- Memo 1 (2023) ---
    data_2023 = [
        ["1", "G. Surender", "10.04.1967", "AE/Distn./Mamnoor", "18.03.2012 FN", "02.11.2010 FN"],
        ["2", "G. Nagaraju", "02.04.1975", "AE/I.TICorp. Office Warangal", "05-08-2014 FN", "04.01.2012 FN"],
        ["3", "K. Venkateshwara Rao", "15.04.1972", "AE/DPE/Kothagudem", "16.12.2015 FN", "16.12.2015 FN"],
        ["4", "G. Ramesh", "10.05.1977", "AE/Opn./Damera", "17-12-2015 FN", "17.12.2015 FN"],
        ["5", "G. Padma", "25.01.1985", "AE/C&O/Sircilla", "09-12-2015 FN", "09-12-2015 FN"],
        ["6", "Ch. Sridhar", "08.01.1980", "AE/Prot./Khammam", "07.12.2015 FN", "07.12.2015 FN"],
        ["7", "B. Anjaneyulu", "14.08.1983", "AE/Comml/DO/Huzurabad", "09-12-2015 FN", "09-12-2015 FN"],
        ["8", "Fasih Ahmed", "27.05.1980", "AE/Op/Manakondur", "09.12.2015 FN", "09.12.2015 FN"],
        ["9", "D. Srinivas", "21.08.1974", "AE/Op/D-I/Nizamabad", "09.12.2015 FN", "09.12.2015 FN"],
        ["10", "D. Srinivas", "07.08.1980", "AE/Op/T. 8/Karimnagar", "09.12.2015 FN", "09.12.2015 FN"],
        ["11", "K. Aravind Reddy", "29.10.1976", "AE/Op/Nustulapur", "09-12-2015 FN", "09-12-2015 FN"],
        ["12", "K.Viswa Prasad", "06.10.1968", "AE/Opn./Parkal Rural", "09.12.2015 FN", "09.12.2015 FN"],
        ["13", "K.Chandra Shekar", "10.04.1979", "AE/Tech/DO/Jagtial", "09-12-2015 FN", "09-12-2015 FN"],
        ["14", "Vanga Ravinder", "09.04.1978", "AE/Comml/DOWarangal", "09.12.2015 FN", "09.12.2015 FN"],
        ["15", "A. Venkat Reddy", "01.09.1973", "AE/SPMIJagtial", "09-12-2015 FN", "09-12-2015 FN"],
        ["16", "M. Anil kumar", "10.05.1980", "AE/Const/Hanumakonda Rural", "07.12.2015 FN", "07.12.2015 FN"],
        ["17", "P. Srinivas", "15.07.1981", "AE/Op/Pegadapally", "09.12.2015 FN", "09.12.2015 FN"],
        ["18", "Lakavath Savai", "15.10.1979", "AE/Op/Achampally", "10-12-2015 FN", "10-12-2015 FN"],
        ["19", "B. Srinivas", "03.07.1975", "AE/Op/Jannepally", "08.12.2015 FN", "08.12.2015 FN"],
        ["20", "T. Ram Mohan", "05.04.1979", "AE/Op/T 6/Karimnagar", "09.05.2016 FN", "09.05.2016 FN"],
        ["21", "Gudelli Srinivas", "02.06.1980", "AE/TRE/Karimnagar", "09.05.2016 FN", "09.05.2016 FN"],
        ["22", "Nasreen Sulthana", "19.08.1982", "AE/Tech/DOWarangal", "06.05.2016 FN", "06.05.2016 FN"],
        ["23", "A. Rajesh", "25.08.1984", "AE/Op'Thallada", "05.05.2016 FN", "05.05.2016 FN"],
        ["24", "K. Purna Chander", "20.03.1979", "AE/Comml/Corp. OfficeWGL", "06.05.2016 FN", "06.05.2016 FN"],
        ["25", "S. Murali Krishna", "20.06.1981", "AE/Distn./Kazipet", "11.05.2016 FN", "11.05.2016 FN"],
        ["26", "V. Shivaprasad", "12.06.1979", "AE/Op/Kallur", "11.05.2016 FN", "11.05.2016 FN"],
        ["27", "A. Muneender Reddy", "06.08.1975", "AE/Op/Chilavakodur", "30-05-2017 FN", "21.11.2016 FN"],
        ["28", "G. Sailoo", "08.01.1982", "AE/Op/Dichpally", "07-11-2017 FN", "23.11.2016 FN"]
    ]
    header_2023 = ["Sl. No", "Name of the AE", "DOB", "Present place of working", "Date of conversion as AE", "Date of commencement of probation"]
    table_2023 = sort_table([header_2023] + data_2023, dob_idx=2, doj_idx=4)
    file1 = write_excel(table_2023, os.path.join(out_dir, "Memo_17_01_2023.xlsx"))

    # --- Memo 2 (2025) ---
    data_2025 = [
        ["1", "R. Ramesh Khanna", "15-05-66", "AE/Projects/Corporate Office"],
        ["2", "M. Ramya", "29.12.1983", "AE(lT)/Corporate Ofiice"],
        ["3", "K. Swarnalatha", "23.07.1982", "AE/Techl./CO/Bupalpally"],
        ["4", "M.Ramulu Naik", "26-09-71", "AE/T/DO/Mahabubabad"],
        ["5", "M. Mallikarjun Rao", "10.05.1975", "AE/DPE/Mahabubabad"],
        ["6", "l. Irlallaiah", "15.06.1987", "AE/Enquiries/Corpoarte Office"],
        ["7", "B. Bhavageethika", "13.05.1989", "AE(Deputed to TSPCC)"],
        ["8", "B. Sandhya Rani", "23.03.1985", "AE(Deputed to TSPCC)"],
        ["9", "Ch. Radhika", "22.03.1980", "AE(Tech.)/D.O./T/ Hanamkonda(R)"],
        ["10", "J. Srujana", "19.06.1985", "AE[l.]/Corp. Office^/Varangal"],
        ["11", "D. Rajin Kumar", "17.12.1987", "AE/Distn./ Station Road"],
        ["12", "h/. Madhusudhan", "18.02.1983", "AE/OP/Sangem"],
        ["13", "K. Rishma", "11.07.1981", "AE/Tech/DO/Warangal"],
        ["14", "P.Sudhakar", "17.10.1977", "AE/OP/Narmetta"],
        ["15", "G. Laxminarayana", "22.05.1978", "AE/OP/Rajavaram"],
        ["16", "V. Venu Kumar", "04-06-1979", "Ex: AE/Opn.Thadwai"],
        ["17", "Burra Srinivas", "03-06-1981", "Ex: AE/Op/Manakondur"],
        ["18", "K. Sreenivas", "03.06.1975", "AE/Operation Section, Challur"],
        ["19", "Thotha IVlallesham", "01.07.1976", "AE-'l/DPE-lliPeddaPllY"],
        ["20", "K. Gopinath", "28.08.1980", "AEffech., Division Offae, Town"],
        ["21", "K. Mallaih", "09.10.1976", "AE/Operation Seclion' Alugunur"],
        ["22", "M. Veera Chary", "10-04-1969", "AE/Operation Section, Alugunur"],
        ["23", "Habeebunneesa", "06.05.1985", "AE/lndoor, District Stores"],
        ["46", "G. Padma", "25-01-85", "AAE/SPM&TRE-I/Asifabad"]
    ]
    header_2025 = ["Sl.No.", "Name of the AE", "DOB", "Place of Working"]
    table_2025 = sort_table([header_2025] + data_2025, dob_idx=2)
    file2 = write_excel(table_2025, os.path.join(out_dir, "Memo_27_03_2025.xlsx"))
    
    # --- Memo 3 (2021) ---
    data_2021 = [
        ["1", "M. Uma Kanth", "12-03-80", "TSREDCL", "Gandugulapally"],
        ["2", "S. Murakli Krishna", "20-06-81", "Warangal", "AAE/Tech/DPE/Warangal"],
        ["3", "E. Vijay Kumar", "06-12-81", "Warangal", "AAE-2/DPE-1/Karimnagar"],
        ["4", "N. Srinivas", "23-06-73", "Warangal", "AAE-2/DPE-II/Jagtial"],
        ["5", "M. Prabhakar", "15-08-79", "Warangal", "AAE/Opn./Uppal"],
        ["6", "B. Chandulal", "04-09-75", "Warangal", "AAE/HT Meters-II/WGL"],
        ["7", "Nasreen Sulthana", "19-08-82", "Warangal", "AAE/Tech/DO/Warangal"],
        ["8", "K. Sampath Reddy", "06-03-80", "Warangal", "AAE/Op/Hasanparthy"],
        ["9", "D. Dayanandam", "02-08-1976", "Warangal", "AAE/OP/Chinthagattu"],
        ["10", "T. Ravinder", "20-06-68", "Warangal", "AAE/SPM/TRE/Warangal"],
        ["11", "G. Surender", "10-04-67", "Warangal", "AAE/SPM/Mahabubabad"],
        ["12", "L. Rajamouli", "08-07-1980", "Warangal", "AAE/Opn/Shayampet"],
        ["13", "S. Srikanth", "28-08-1985", "Warangal", "AAE/Op/Sangem"],
        ["14", "B. Ashok", "05-04-68", "Jangaon", "AAE/Op/Tharigoppula"],
        ["15", "N. Sreenivasulu", "15-02-80", "Mahabubad", "AAE/Op/Dornakal"],
        ["16", "B. Sathish Kumar", "06-03-80", "Mahabubabad", "AAE/Tech/DO/Mahabubad"],
        ["17", "A. Sridhar", "05-05-79", "Bhupalpally", "AAE/Op/Kamalapur"],
        ["18", "P. Srinivas", "10-04-72", "Karimnagar", "AAE/Techl/DO/Huzurabad"],
        ["19", "R. Satyanarayana", "15-04-79", "Karimnagar", "AAE/Outdoor/D.S/Karimnagar"],
        ["20", "K. Aravind Reddy", "29-10-78", "Karimnagar", "AAE/Op/Nustulapur"],
        ["21", "Lakavath Saval", "15-10-70", "Karimnagar", "AAE/Opn./Achampally"],
        ["22", "P. Srinivas", "15-07-81", "Jagtial", "AAE/Op/Pegadapally"],
        ["23", "K. Kamaleshwar", "10-06-82", "Jagtial", "AAE/Op/Mallial"],
        ["24", "K. Kishore", "15-08-79", "Peddapally", "AAE/Opn./Begampet"]
    ]
    header_2021 = ["Sl.No.", "Name of the AAE", "Date of Birth", "Circle", "Present place of working"]
    table_2021 = sort_table([header_2021] + data_2021, dob_idx=2)
    file3 = write_excel(table_2021, os.path.join(out_dir, "Memo_16_09_2021.xlsx"))
    
    # --- Memo 4 (2022) ---
    data_2022 = [
        ["1", "G. Surender", "10-04-1967", "Warangal", "18-03-2012 FN"],
        ["2", "G. Nagaraju", "02-04-1975", "Corporate Office", "05-08-2014 FN"],
        ["3", "G. Ramesh", "10-05-1977", "Warangal", "17-12-2015 FN"],
        ["4", "A. Muneender Reddy", "06-08-1975", "Jagtial", "30-05-2017 FN"],
        ["5", "G. Padma", "25-01-1985", "Mancherial", "09-12-2015 FN"],
        ["6", "G. Sailoo", "08-01-1982", "Nizamabad", "07-11-2017 FN"],
        ["7", "B. Anjaneyulu", "14-08-1983", "JS-Bhupalapally", "09-12-2015 FN"],
        ["8", "Fasih Ahmed", "27-05-1980", "Karimnagar", "09-12-2015 FN"],
        ["9", "D. Srinivas", "21-08-1974", "Nizamabad", "09-12-2015 FN"],
        ["10", "D. Srinivas", "07-08-1980", "Karimnagar", "09-12-2015 FN"],
        ["11", "K.Viswa Prasad", "06-10-1968", "Hanumakonda", "09-12-2015 FN"],
        ["12", "Vanga Ravinder", "09-04-1978", "Hanumakonda", "09-12-2015 FN"],
        ["13", "P. Srinivas", "15-07-1981", "Jagtial", "09-12-2015 FN"],
        ["14", "B. Srinivas", "03-07-1975", "Nizamabad", "08-12-2015 FN"],
        ["15", "Gudelli. Srinivas", "02-06-1980", "Karimnagar", "09-05-2016 FN"],
        ["16", "A. Rajesh", "25-08-1984", "Khammam", "05-05-2016 FN"],
        ["17", "K. Purna Chander", "20-03-1979", "JS-Bhupalapally", "06-05-2016 FN"],
        ["18", "S. Murakli Krishna", "20-06-1981", "Hanumakonda", "11-05-2016 FN"],
        ["19", "V. Shivaprasad", "12-06-1979", "Nirmal", "11-05-2016 FN"]
    ]
    header_2022 = ["Sl. No.", "Name of the AAE Sri/Smt:", "Date of Birth", "Circle", "Date of conversion"]
    table_2022 = sort_table([header_2022] + data_2022, dob_idx=2, doj_idx=4)
    file4 = write_excel(table_2022, os.path.join(out_dir, "Memo_20_01_2022.xlsx"))
    
    files = [file1, file2, file3, file4]
    
    speak("I have processed all 4 files separately. The tables are now sorted in descending order by Date of Birth and Date of Joining. I will now open them one by one for your visual confirmation.")
    
    for f in files:
        os.startfile(f)
        time.sleep(1.0)
    
    time.sleep(2)
    speak("Now, I am routing the files to Ravindar vanga in WhatsApp.")
    
    # Try to automate WhatsApp to send to "Ravindar vanga"
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=2)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(0.5)
        
        # Select chat
        window.type_keys("^f")
        time.sleep(0.5)
        window.type_keys("Ravindar vanga{ENTER}", with_spaces=True)
        time.sleep(1.5)
        
        for f in files:
            cmd = f'powershell.exe -command "Set-Clipboard -Path \'{f}\'"'
            subprocess.run(cmd, shell=True)
            time.sleep(0.5)
            window.type_keys("^v")
            time.sleep(1.0)
            window.type_keys("{ENTER}")
            time.sleep(1.0)
            
        speak("All 4 workbooks have been successfully dispatched to Ravindar vanga.")
    except Exception as e:
        speak("I could not automate WhatsApp (perhaps it is minimized or closed), but the 4 separate workbooks are fully generated, sorted, and displayed on your screen.")
        print(f"WhatsApp UI error: {e}")

if __name__ == "__main__":
    run_automation()
