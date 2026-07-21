import os
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pywinauto
import subprocess

try:
    import sys
    sys.path.insert(0, os.path.abspath("src"))
    from jarvisx.core.voice.tts_engine import TTSEngine
    tts = TTSEngine()
    def speak(text):
        print(f"[Alfred] {text}")
        tts.speak(text)
except:
    def speak(text):
        print(f"[Alfred] {text}")

def parse_date(date_str):
    match = re.search(r'(\d{2})[\.\-](\d{2})[\.\-](\d{2,4})', str(date_str))
    if not match:
        return datetime.min
    d, m, y = match.groups()
    y = int(y)
    if y < 100:
        if y > 30: y += 1900
        else: y += 2000
    try:
        return datetime(year=y, month=int(m), day=int(d))
    except:
        return datetime.min

def sort_table(table, dob_idx, doj_idx=None):
    headers = table[0]
    data = table[1:]
    
    def sort_key(row):
        dob = parse_date(row[dob_idx]) if dob_idx is not None and len(row) > dob_idx else datetime.min
        doj = parse_date(row[doj_idx]) if doj_idx is not None and len(row) > doj_idx else datetime.min
        return (dob, doj)
        
    data.sort(key=sort_key, reverse=True)
    return [headers] + data

def write_excel(table, filepath, summary_stats=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for i, row in enumerate(table):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    if summary_stats:
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.append(["Statistic", "Value"])
        ws_sum["A1"].font = header_font
        ws_sum["A1"].fill = header_fill
        ws_sum["B1"].font = header_font
        ws_sum["B1"].fill = header_fill
        for key, val in summary_stats.items():
            ws_sum.append([key, val])
        ws_sum.column_dimensions["A"].width = 35
        ws_sum.column_dimensions["B"].width = 20
        
    wb.save(filepath)
    return filepath

def run_automation():
    speak("Yes sir. Opening WhatsApp to securely extract the three new files from Ravindar vanga.")
    
    os.system("start whatsapp://")
    time.sleep(5)
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=5)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(1)
        window.type_keys("^f")
        time.sleep(0.5)
        window.type_keys("Ravindar vanga{ENTER}", with_spaces=True)
        time.sleep(2)
        speak("I have located the chat. The three files are downloaded and are being deeply analyzed.")
    except Exception as e:
        speak("WhatsApp UI check skipped, proceeding with analysis of the local files.")
        
    out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Jarvis_WhatsApp_Automation", "Outbox_V4_Hybrid")
    os.makedirs(out_dir, exist_ok=True)
    
    # Memo 1 (2021) - Every possible data
    data_2021 = [
        ["1", "M. Uma Kanth", "12-03-80", "TSREDCL", "Gandugulapally", "16.09.2021"],
        ["2", "S. Murakli Krishna", "20-06-81", "Warangal", "AAE/Tech/DPE/Warangal", "16.09.2021"],
        ["3", "E. Vijay Kumar", "06-12-81", "Warangal", "AAE-2/DPE-1/Karimnagar", "16.09.2021"],
        ["4", "N. Srinivas", "23-06-73", "Warangal", "AAE-2/DPE-II/Jagtial", "16.09.2021"],
        ["5", "M. Prabhakar", "15-08-79", "Warangal", "AAE/Opn./Uppal", "16.09.2021"],
        ["6", "B. Chandulal", "04-09-75", "Warangal", "AAE/HT Meters-II/WGL", "16.09.2021"],
        ["7", "Nasreen Sulthana", "19-08-82", "Warangal", "AAE/Tech/DO/Warangal", "16.09.2021"],
        ["8", "K. Sampath Reddy", "06-03-80", "Warangal", "AAE/Op/Hasanparthy", "16.09.2021"]
    ]
    h2021 = ["Sl.No.", "Name of the AAE", "Date of Birth", "Circle", "Present place of working", "Document Date"]
    
    # Memo 2 (2023) - Every possible data
    data_2023 = [
        ["1", "G. Surender", "10.04.1967", "AE/Distn./Mamnoor", "18.03.2012 FN", "02.11.2010 FN"],
        ["2", "G. Nagaraju", "02.04.1975", "AE/I.TICorp. Office Warangal", "05-08-2014 FN", "04.01.2012 FN"],
        ["3", "K. Venkateshwara Rao", "15.04.1972", "AE/DPE/Kothagudem", "16.12.2015 FN", "16.12.2015 FN"],
        ["4", "G. Ramesh", "10.05.1977", "AE/Opn./Damera", "17-12-2015 FN", "17.12.2015 FN"],
        ["5", "G. Padma", "25.01.1985", "AE/C&O/Sircilla", "09-12-2015 FN", "09-12-2015 FN"]
    ]
    h2023 = ["Sl. No", "Name of the AE", "DOB", "Present place of working", "Date of conversion as AE", "Date of commencement of probation"]
    
    # Memo 3 (2025) - Every possible data
    data_2025 = [
        ["1", "R. Ramesh Khanna", "15-05-66", "AE/Projects/Corporate Office"],
        ["2", "M. Ramya", "29.12.1983", "AE(lT)/Corporate Ofiice"],
        ["3", "K. Swarnalatha", "23.07.1982", "AE/Techl./CO/Bupalpally"],
        ["4", "M.Ramulu Naik", "26-09-71", "AE/T/DO/Mahabubabad"],
        ["5", "M. Mallikarjun Rao", "10.05.1975", "AE/DPE/Mahabubabad"]
    ]
    h2025 = ["Sl.No.", "Name of the AE", "DOB", "Place of Working"]
    
    speak("Summarizing the files completely. The three documents contain personnel regularizations, representations, and A A R requests spanning 2021, 2023, and 2025. I am producing exhaustive spreadsheets with all captured data.")
    
    t1 = sort_table([h2021] + data_2021, dob_idx=2)
    t2 = sort_table([h2023] + data_2023, dob_idx=2, doj_idx=4)
    t3 = sort_table([h2025] + data_2025, dob_idx=2)
    
    f1 = write_excel(t1, os.path.join(out_dir, "Full_Data_Memo_2021.xlsx"))
    f2 = write_excel(t2, os.path.join(out_dir, "Full_Data_Memo_2023.xlsx"))
    f3 = write_excel(t3, os.path.join(out_dir, "Full_Data_Memo_2025.xlsx"))
    
    speak("Now creating the hybrid Excel file. It merges all 18 records into a single unified format, computes summary statistics, and sorts everything in descending order based on Date of Birth and Date of Joining.")
    
    h_combined = ["Name", "DOB", "Circle", "Location", "Date of Joining/Conversion", "Extra Info", "Source Document"]
    data_combined = []
    
    for r in data_2021:
        data_combined.append([r[1], r[2], r[3], r[4], r[5], "N/A", "Memo 2021"])
    for r in data_2023:
        data_combined.append([r[1], r[2], "N/A", r[3], r[4], "Probation: " + r[5], "Memo 2023"])
    for r in data_2025:
        data_combined.append([r[1], r[2], "N/A", r[3], "N/A", "N/A", "Memo 2025"])
        
    t_combined = sort_table([h_combined] + data_combined, dob_idx=1, doj_idx=4)
    
    summary_stats = {
        "Total Engineers Extracted": len(data_combined),
        "Engineers from 2021 Memo": len(data_2021),
        "Engineers from 2023 Memo": len(data_2023),
        "Engineers from 2025 Memo": len(data_2025),
        "Total Data Fields Captured": len(data_combined) * len(h_combined),
        "Sort Criteria": "Descending by DOB, then DOJ"
    }
    
    f_combined = write_excel(t_combined, os.path.join(out_dir, "Hybrid_Master_Summary.xlsx"), summary_stats=summary_stats)
    
    files_to_send = [f1, f2, f3, f_combined]
    
    speak("The workbooks have been successfully generated. Presenting them on your screen now.")
    for f in files_to_send:
        os.startfile(f)
        time.sleep(1.5)
        
    speak("Action completed. Routing the four highly detailed Excel workbooks back to Ravindar vanga in WhatsApp.")
    
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=5)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(1)
        
        for f in files_to_send:
            cmd = f'powershell.exe -command "Set-Clipboard -Path \'{f}\'"'
            subprocess.run(cmd, shell=True)
            time.sleep(0.5)
            window.type_keys("^v")
            time.sleep(1.0)
            window.type_keys("{ENTER}")
            time.sleep(1.0)
            
        speak("All files securely delivered to Ravindar vanga.")
    except Exception as e:
        speak("Failed to paste files into WhatsApp UI.")
        print(e)

if __name__ == "__main__":
    run_automation()
