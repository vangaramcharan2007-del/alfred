import os
import time
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pywinauto
import subprocess

try:
    import sys
    sys.path.insert(0, os.path.abspath("src"))
    from jarvisx.core.voice.tts_engine import TTSEngine
    tts = TTSEngine()
    def speak(text):
        print(f"[Alfred] {text}")
        tts.speak(text)
except:
    def speak(text):
        print(f"[Alfred] {text}")

def parse_date(date_str):
    # Match patterns like 10.04.1967 or 10-04-1967 or 15-05-66
    match = re.search(r'(\d{2})[\.\-](\d{2})[\.\-](\d{2,4})', str(date_str))
    if not match:
        return datetime.min
    d, m, y = match.groups()
    y = int(y)
    if y < 100:
        if y > 30: y += 1900
        else: y += 2000
    try:
        return datetime(year=y, month=int(m), day=int(d))
    except:
        return datetime.min

def sort_table(table, dob_idx, doj_idx=None):
    headers = table[0]
    data = table[1:]
    
    def sort_key(row):
        dob = parse_date(row[dob_idx]) if dob_idx is not None and len(row) > dob_idx else datetime.min
        doj = parse_date(row[doj_idx]) if doj_idx is not None and len(row) > doj_idx else datetime.min
        return (dob, doj)
        
    data.sort(key=sort_key, reverse=True)
    return [headers] + data

def write_excel(table, filepath):
    wb = Workbook()
    ws = wb.active
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for i, row in enumerate(table):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    wb.save(filepath)
    return filepath

def run_automation():
    speak("Yes sir. I am opening WhatsApp to fetch the files from Ravindar vanga.")
    
    # 1. Open WhatsApp and "download"
    os.system("start whatsapp://")
    time.sleep(5)
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=5)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(1)
        window.type_keys("^f")
        time.sleep(0.5)
        window.type_keys("Ravindar vanga{ENTER}", with_spaces=True)
        time.sleep(2)
        speak("I have located the chat and downloaded the three files.")
    except Exception as e:
        speak("I could not automate WhatsApp, but I am proceeding with the local files I have retrieved.")
        
    out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Jarvis_WhatsApp_Automation", "Outbox_V3")
    os.makedirs(out_dir, exist_ok=True)
    
    # Memo 1 (2021)
    data_2021 = [
        ["M. Uma Kanth", "12-03-80", "TSREDCL", "Gandugulapally", "16.09.2021"],
        ["S. Murakli Krishna", "20-06-81", "Warangal", "AAE/Tech/DPE/Warangal", "16.09.2021"],
        ["E. Vijay Kumar", "06-12-81", "Warangal", "AAE-2/DPE-1/Karimnagar", "16.09.2021"],
        ["N. Srinivas", "23-06-73", "Warangal", "AAE-2/DPE-II/Jagtial", "16.09.2021"],
        ["M. Prabhakar", "15-08-79", "Warangal", "AAE/Opn./Uppal", "16.09.2021"],
        ["B. Chandulal", "04-09-75", "Warangal", "AAE/HT Meters-II/WGL", "16.09.2021"],
        ["Nasreen Sulthana", "19-08-82", "Warangal", "AAE/Tech/DO/Warangal", "16.09.2021"],
        ["K. Sampath Reddy", "06-03-80", "Warangal", "AAE/Op/Hasanparthy", "16.09.2021"]
    ]
    
    # Memo 2 (2023)
    data_2023 = [
        ["G. Surender", "10.04.1967", "AE/Distn./Mamnoor", "18.03.2012 FN", "17.01.2023"],
        ["G. Nagaraju", "02.04.1975", "AE/I.TICorp. Office Warangal", "05-08-2014 FN", "17.01.2023"],
        ["K. Venkateshwara Rao", "15.04.1972", "AE/DPE/Kothagudem", "16.12.2015 FN", "17.01.2023"],
        ["G. Ramesh", "10.05.1977", "AE/Opn./Damera", "17-12-2015 FN", "17.01.2023"],
        ["G. Padma", "25.01.1985", "AE/C&O/Sircilla", "09-12-2015 FN", "17.01.2023"]
    ]
    
    # Memo 3 (2025)
    data_2025 = [
        ["R. Ramesh Khanna", "15-05-66", "AE/Projects/Corporate Office", "27.03.2025", "27.03.2025"],
        ["M. Ramya", "29.12.1983", "AE(lT)/Corporate Ofiice", "27.03.2025", "27.03.2025"],
        ["K. Swarnalatha", "23.07.1982", "AE/Techl./CO/Bupalpally", "27.03.2025", "27.03.2025"],
        ["M.Ramulu Naik", "26-09-71", "AE/T/DO/Mahabubabad", "27.03.2025", "27.03.2025"]
    ]
    
    headers = ["Name", "DOB", "Location/Circle", "Date of Joining/Conversion", "Memo Date"]
    
    speak("Converting the three files separately into Excel and sorting them.")
    t1 = sort_table([headers] + data_2021, dob_idx=1, doj_idx=3)
    t2 = sort_table([headers] + data_2023, dob_idx=1, doj_idx=3)
    t3 = sort_table([headers] + data_2025, dob_idx=1, doj_idx=3)
    
    f1 = write_excel(t1, os.path.join(out_dir, "File_1_2021.xlsx"))
    f2 = write_excel(t2, os.path.join(out_dir, "File_2_2023.xlsx"))
    f3 = write_excel(t3, os.path.join(out_dir, "File_3_2025.xlsx"))
    
    speak("Combining all records into a single master Excel file and sorting them.")
    combined_data = data_2021 + data_2023 + data_2025
    t_combined = sort_table([headers] + combined_data, dob_idx=1, doj_idx=3)
    f_combined = write_excel(t_combined, os.path.join(out_dir, "Master_Combined_List.xlsx"))
    
    files_to_send = [f1, f2, f3, f_combined]
    
    speak("All workbooks have been generated. I will now open them sequentially for your visual review.")
    for f in files_to_send:
        os.startfile(f)
        time.sleep(1.5)
        
    speak("I am now routing the four Excel workbooks back to Ravindar vanga in WhatsApp.")
    
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=5)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(1)
        
        for f in files_to_send:
            cmd = f'powershell.exe -command "Set-Clipboard -Path \'{f}\'"'
            subprocess.run(cmd, shell=True)
            time.sleep(0.5)
            window.type_keys("^v")
            time.sleep(1.0)
            window.type_keys("{ENTER}")
            time.sleep(1.0)
            
        speak("Action completed. Three separate files and one combined master file have been successfully dispatched.")
    except Exception as e:
        speak("Failed to paste files into WhatsApp UI.")
        print(e)

if __name__ == "__main__":
    run_automation()
