import os
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import pywinauto
import subprocess

try:
    import sys
    sys.path.insert(0, os.path.abspath("src"))
    from jarvisx.core.voice.tts_engine import TTSEngine
    tts = TTSEngine()
    def speak(text):
        print(f"[Alfred] {text}")
        tts.speak(text)
except:
    def speak(text):
        print(f"[Alfred] {text}")

def write_excel(table, filepath, summary_stats=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Names Only"
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for i, row in enumerate(table):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2
        
    if summary_stats:
        ws_sum = wb.create_sheet(title="Summary")
        ws_sum.append(["Statistic", "Value"])
        ws_sum["A1"].font = header_font
        ws_sum["A1"].fill = header_fill
        ws_sum["B1"].font = header_font
        ws_sum["B1"].fill = header_fill
        for key, val in summary_stats.items():
            ws_sum.append([key, val])
        ws_sum.column_dimensions["A"].width = 35
        ws_sum.column_dimensions["B"].width = 20
        
    wb.save(filepath)
    return filepath

def run_automation():
    speak("Yes sir. I am isolating exactly just the names found exclusively within the tables. I will ignore all letterheads, signatures, and extra columns, providing you with a pure, summarized list of names.")
    
    out_dir = os.path.join(os.path.expanduser("~"), "Desktop", "Jarvis_WhatsApp_Automation", "Outbox_V6_Names_Only")
    os.makedirs(out_dir, exist_ok=True)
    
    # Memo 2021 (46 Engineers)
    data_2021_names = [
        ["M. Uma Kanth"], ["S. Murakli Krishna"], ["E. Vijay Kumar"], ["N. Srinivas"], ["M. Prabhakar"], 
        ["B. Chandulal"], ["Nasreen Sulthana"], ["K. Sampath Reddy"], ["D. Dayanandam"], ["T. Ravinder"], 
        ["G. Surender"], ["L. Rajamouli"], ["S. Srikanth"], ["B. Ashok"], ["N. Sreenivasulu"], 
        ["B. Sathish Kumar"], ["A. Sridhar"], ["P. Srinivas"], ["R. Satyanarayana"], ["K. Aravind Reddy"], 
        ["Lakavath Saval"], ["P. Srinivas"], ["K. Kamaleshwar"], ["K. Kishore"], ["K.V. Koteshwar Rao"], 
        ["R. Suresh"], ["A. Balaji"], ["A. Mamatha"], ["G. Sailoo"], ["B. Srinivas"], ["R. Sumitha"], 
        ["N. Nageshwar Rao"], ["S. Venkata Ramana Chary"], ["K. Jyothi"], ["D. Arpitha"], ["Krishnapuram Ramesh"], 
        ["Ch. Thirupathi Reddy"], ["N. Bhumaiah"], ["M. Gopi Krishna"], ["B. Krishna Rao"], ["T. Srinivas"], 
        ["D. Rajender"], ["V. Shivaprasad"], ["MD. Azimuddin"], ["D. Balaiah"], ["G. Padma"]
    ]
    
    # Memo 2023 (28 Engineers)
    data_2023_names = [
        ["G. Surender"], ["G. Nagaraju"], ["K. Venkateshwara Rao"], ["G. Ramesh"], ["G. Padma"], 
        ["Ch. Sridhar"], ["B. Anjaneyulu"], ["Fasih Ahmed"], ["D. Srinivas"], ["D. Srinivas"], 
        ["K. Aravind Reddy"], ["K.Viswa Prasad"], ["K.Chandra Shekar"], ["Vanga Ravinder"], ["A. Venkat Reddy"], 
        ["M. Anil kumar"], ["P. Srinivas"], ["Lakavath Savai"], ["B. Srinivas"], ["T. Ram Mohan"], 
        ["Gudelli Srinivas"], ["Nasreen Sulthana"], ["A. Rajesh"], ["K. Purna Chander"], ["S. Murali Krishna"], 
        ["V. Shivaprasad"], ["A. Muneender Reddy"], ["G. Sailoo"]
    ]
    
    # Memo 2025 (92 Engineers)
    data_2025_names = [
        ["R. Ramesh Khanna"], ["M. Ramya"], ["K. Swarnalatha"], ["M.Ramulu Naik"], ["M. Mallikarjun Rao"], 
        ["l. Irlallaiah"], ["B. Bhavageethika"], ["B. Sandhya Rani"], ["Ch. Radhika"], ["J. Srujana"], 
        ["D. Rajin Kumar"], ["h/. Madhusudhan"], ["K. Rishma"], ["P.Sudhakar"], ["G. Laxminarayana"], 
        ["V. Venu Kumar"], ["Burra Srinivas"], ["K. Sreenivas"], ["Thotha IVlallesham"], ["K. Gopinath"], 
        ["K. Mallaih"], ["M. Veera Chary"], ["Habeebunneesa"], ["P. Ravi"], ["A. Bhoomeshwar"], 
        ["D. Ravi"], ["V. Gangadhar"], ["R. Raghunath"], ["B. Pradeep Kumar"], ["C. Anil Kumar"], 
        ["P. Rama Swamy"], ["B.Sridhar"], ["S. Rama Rao"], ["P. Venkat"], ["K. Anil Kumar"], 
        ["D. Uma Maheswara Rao"], ["N. Srinivasa Rao"], ["K. Rama Rao"], ["S. Rami Reddy"], ["M.Srinivasa Rao"], 
        ["R. Bhaskar Rao"], ["J. Anil Kumar"], ["D. Ashok"], ["J. Ramya"], ["A. Umamaheswari"], 
        ["Azmeera lndira"], ["S. Swathi"], ["M. Lakshmi Narasimha Rao"], ["G.Ravinder"], ["P.Madhu Babu"], 
        ["B. Uma Rao"], ["R. Manasa"], ["V. Durgaprasad"], ["P. Vinod"], ["C. Naveen Reddy"], 
        ["K. Nageshwar"], ["A. Aravind"], ["Boini. Laxmi Narsimha"], ["M. Sathish Reddy"], ["Ch. Laxman Rao"], 
        ["lrfan Ahmed"], ["Dussa Srinivas"], ["P. Sunder Kumar"], ["G. Surender"], ["Regalla Rajesh"], 
        ["Badela Ramesh"], ["Banka Nagalaxmi"], ["T.Ravinder"], ["N.Venugopal"], ["C.Thirupathi"], 
        ["G.Chungulal Naik"], ["P.Janardhan"], ["V.Ashok"], ["J.Amarender"], ["Ch.Mallaiah"], 
        ["N.Manohar"], ["B. Yohan"], ["K.Siddiramulu"], ["Md.Zakeer Ali"], ["V.Venkat Narayana"], 
        ["D.Durgasiva Prasad"], ["Md. Maqdoom Ali"], ["V.Prem Kumar"], ["K. Srinivas Rao"], ["K. Raja Ratnam"], 
        ["V. Sudhakar Reddy"], ["K. Anand"], ["D. Venkateshwar Rao"], ["A. Raghu Ramaiah"], ["M. lndrasena"], 
        ["M. Ramesh"], ["G. Satyarayana"]
    ]
    
    speak("I have filtered out all letterheads, dates, and locations. I am generating three separate Excel sheets containing exclusively the isolated names.")
    
    f1 = write_excel([["Name of Engineer (2021)"]] + data_2021_names, os.path.join(out_dir, "Names_Only_2021.xlsx"))
    f2 = write_excel([["Name of Engineer (2023)"]] + data_2023_names, os.path.join(out_dir, "Names_Only_2023.xlsx"))
    f3 = write_excel([["Name of Engineer (2025)"]] + data_2025_names, os.path.join(out_dir, "Names_Only_2025.xlsx"))
    
    speak("I am now generating the Hybrid Summary file which combines all 166 table-bound names into a single unified list.")
    
    combined_names = data_2021_names + data_2023_names + data_2025_names
    
    # Sort alphabetically to keep it organized since we dropped dates
    combined_names.sort(key=lambda x: x[0])
    
    summary_stats = {
        "Total Names Extracted": len(combined_names),
        "Source Documents": "16.09.2021, 17.01.2023, 27.03.2025",
        "Filtration Method": "Strict Table Extraction (Non-tabular names dropped)",
        "Sort Order": "Alphabetical (A-Z)"
    }
    
    f_combined = write_excel([["Engineer Name (Hybrid Master List)"]] + combined_names, os.path.join(out_dir, "Ultimate_Names_Summary.xlsx"), summary_stats=summary_stats)
    
    files_to_send = [f1, f2, f3, f_combined]
    
    speak("The stripped-down Name-Only spreadsheets have been generated. Presenting them on your screen now.")
    for f in files_to_send:
        os.startfile(f)
        time.sleep(1.5)
        
    speak("Routing these four highly specific files directly to Ravindar vanga in WhatsApp.")
    
    # Automate WhatsApp
    try:
        app = pywinauto.Application(backend="uia").connect(title_re=".*WhatsApp.*", timeout=5)
        window = app.window(title_re=".*WhatsApp.*")
        window.set_focus()
        time.sleep(1)
        
        for f in files_to_send:
            cmd = f'powershell.exe -command "Set-Clipboard -Path \'{f}\'"'
            subprocess.run(cmd, shell=True)
            time.sleep(0.5)
            window.type_keys("^v")
            time.sleep(1.0)
            window.type_keys("{ENTER}")
            time.sleep(1.0)
            
        speak("All Name-Only files securely delivered to Ravindar vanga.")
    except Exception as e:
        speak("Failed to paste files into WhatsApp UI.")
        print(e)

if __name__ == "__main__":
    run_automation()
