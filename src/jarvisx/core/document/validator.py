import os
from openpyxl import load_workbook

class ExcelValidator:
    """Validates the output Excel workbook generated from a document."""

    @staticmethod
    def validate(filepath: str) -> dict:
        """
        Validates the Excel file.
        Checks: workbook created successfully, correct sheet count, row count,
        column count, empty tables detected, corrupted docs.
        """
        result = {
            "success": False,
            "corrupted": False,
            "empty": False,
            "sheet_count": 0,
            "total_rows": 0,
            "total_cols": 0,
            "errors": []
        }

        if not os.path.exists(filepath):
            result["errors"].append(f"File does not exist: {filepath}")
            return result

        try:
            wb = load_workbook(filepath, read_only=True)
            result["sheet_count"] = len(wb.sheetnames)
            
            if result["sheet_count"] == 0:
                result["empty"] = True
                result["errors"].append("Workbook contains no sheets.")
                return result

            # Validate each sheet
            total_r = 0
            total_c = 0
            has_data = False
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                max_r = ws.max_row
                max_c = ws.max_column
                
                total_r += max_r
                total_c += max_c
                
                if max_r > 0 and max_c > 0:
                    has_data = True

            result["total_rows"] = total_r
            result["total_cols"] = total_c

            if not has_data:
                result["empty"] = True
                result["errors"].append("Workbook contains no data (empty tables).")
            else:
                result["success"] = True

        except Exception as e:
            result["corrupted"] = True
            result["errors"].append(f"Workbook is corrupted or invalid: {str(e)}")

        return result
