import os
import docx
from openpyxl import Workbook
from typing import Dict, Any

from jarvisx.core.document.converter_base import DocumentConverterBase
from jarvisx.core.document.validator import ExcelValidator

class WordToExcelConverter(DocumentConverterBase):
    """Converts tables from Word (.docx) documents into Excel (.xlsx) workbooks."""

    def convert(self, source_path: str, destination_path: str) -> Dict[str, Any]:
        result = {
            "success": False,
            "metrics": {
                "tables_found": 0,
                "rows_processed": 0,
                "cells_processed": 0
            },
            "error": None
        }

        try:
            # 1. Load Document
            doc = docx.Document(source_path)
            tables = doc.tables
            result["metrics"]["tables_found"] = len(tables)

            if len(tables) == 0:
                result["error"] = "No tables found in document."
                return result

            # 2. Create Workbook
            wb = Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # 3. Extract Tables
            for i, table in enumerate(tables):
                sheet_name = f"Table_{i+1}"
                ws = wb.create_sheet(title=sheet_name)

                for r_idx, row in enumerate(table.rows):
                    row_data = []
                    for cell in row.cells:
                        # Extract raw text, replacing common weird unicode chars if necessary
                        text = cell.text.strip()
                        row_data.append(text)
                        result["metrics"]["cells_processed"] += 1
                    
                    ws.append(row_data)
                    result["metrics"]["rows_processed"] += 1

            # 4. Save Workbook
            wb.save(destination_path)
            
            # 5. Validate Output
            val_res = ExcelValidator.validate(destination_path)
            if val_res["success"]:
                result["success"] = True
            else:
                result["error"] = "Validation failed: " + "; ".join(val_res["errors"])

        except Exception as e:
            result["error"] = str(e)

        return result

    def validate(self, destination_path: str) -> bool:
        res = ExcelValidator.validate(destination_path)
        return res["success"]
